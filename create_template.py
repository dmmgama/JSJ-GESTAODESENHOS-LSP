"""
Generate minimal LPP_TEMPLATE.xlsx if it doesn't exist.
Run this once to create the template structure.
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pathlib import Path


def create_lpp_template(output_path: str = "data/LPP_TEMPLATE.xlsx"):
    """
    Create minimal LPP template Excel file with example structure.
    """
    wb = Workbook()
    sheet = wb.active
    sheet.title = "LPP"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    tipo_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    tipo_font = Font(bold=True, size=11)
    
    elemento_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    elemento_font = Font(bold=True, size=10)
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Column widths
    sheet.column_dimensions['A'].width = 15  # Nº.
    sheet.column_dimensions['B'].width = 40  # DESIGNAÇÃO
    sheet.column_dimensions['C'].width = 30  # FICHEIRO
    sheet.column_dimensions['D'].width = 8   # Rev
    sheet.column_dimensions['E'].width = 15  # DATA
    sheet.column_dimensions['F'].width = 12  # ROW_KIND (hidden)
    sheet.column_dimensions['G'].width = 20  # TIPO_KEY (hidden)
    sheet.column_dimensions['H'].width = 15  # ELEMENTO_KEY (hidden)
    
    # Hide technical columns
    sheet.column_dimensions['F'].hidden = True
    sheet.column_dimensions['G'].hidden = True
    sheet.column_dimensions['H'].hidden = True
    
    # Header row (row 1)
    headers = ["Nº.", "DESIGNAÇÃO", "FICHEIRO", "Rev", "DATA", "ROW_KIND", "TIPO_KEY", "ELEMENTO_KEY"]
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(1, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
    
    # Example: TIPO row (row 2)
    sheet.cell(2, 1).value = "BETÃO ARMADO"
    sheet.cell(2, 1).fill = tipo_fill
    sheet.cell(2, 1).font = tipo_font
    sheet.cell(2, 1).border = border_thin
    sheet.cell(2, 6).value = "TIPO"
    sheet.cell(2, 7).value = "BETAO_ARMADO"
    
    # Merge cells for TIPO header
    sheet.merge_cells('A2:H2')
    
    # Example: ELEMENTO row (row 3) - FUN
    sheet.cell(3, 1).value = "FUN - FUNDAÇÕES"
    sheet.cell(3, 1).fill = elemento_fill
    sheet.cell(3, 1).font = elemento_font
    sheet.cell(3, 1).border = border_thin
    sheet.cell(3, 6).value = "ELEMENTO"
    sheet.cell(3, 7).value = "BETAO_ARMADO"
    sheet.cell(3, 8).value = "FUN"
    
    # Merge cells for ELEMENTO header
    sheet.merge_cells('A3:H3')
    
    # Example: DESENHO rows will be inserted here by the app (below row 3)
    # Add placeholder comment
    sheet.cell(4, 1).value = "(Desenhos serão inseridos aqui pela app)"
    sheet.cell(4, 1).font = Font(italic=True, color="808080")
    
    # Example: Another ELEMENTO row (row 5) - PIL
    sheet.cell(6, 1).value = "PIL - PILARES"
    sheet.cell(6, 1).fill = elemento_fill
    sheet.cell(6, 1).font = elemento_font
    sheet.cell(6, 1).border = border_thin
    sheet.cell(6, 6).value = "ELEMENTO"
    sheet.cell(6, 7).value = "BETAO_ARMADO"
    sheet.cell(6, 8).value = "PIL"
    
    sheet.merge_cells('A6:H6')
    
    # Placeholder for PIL drawings
    sheet.cell(7, 1).value = "(Desenhos PIL serão inseridos aqui)"
    sheet.cell(7, 1).font = Font(italic=True, color="808080")
    
    # Example: Another TIPO - ESTRUTURA METÁLICA (row 9)
    sheet.cell(9, 1).value = "ESTRUTURA METÁLICA"
    sheet.cell(9, 1).fill = tipo_fill
    sheet.cell(9, 1).font = tipo_font
    sheet.cell(9, 1).border = border_thin
    sheet.cell(9, 6).value = "TIPO"
    sheet.cell(9, 7).value = "ESTRUTURA_METALICA"
    
    sheet.merge_cells('A9:H9')
    
    # ELEMENTO for ESTRUTURA METÁLICA (row 10)
    sheet.cell(10, 1).value = "VIG - VIGAS"
    sheet.cell(10, 1).fill = elemento_fill
    sheet.cell(10, 1).font = elemento_font
    sheet.cell(10, 1).border = border_thin
    sheet.cell(10, 6).value = "ELEMENTO"
    sheet.cell(10, 7).value = "ESTRUTURA_METALICA"
    sheet.cell(10, 8).value = "VIG"
    
    sheet.merge_cells('A10:H10')
    
    # Placeholder
    sheet.cell(11, 1).value = "(Desenhos VIG serão inseridos aqui)"
    sheet.cell(11, 1).font = Font(italic=True, color="808080")
    
    # Freeze first row
    sheet.freeze_panes = "A2"
    
    # Save
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Template criado: {output_path}")


if __name__ == "__main__":
    create_lpp_template()
