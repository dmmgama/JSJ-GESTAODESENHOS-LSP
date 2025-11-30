"""
LPP Builder - generates/updates LPP.xlsx from database using template.
"""
from pathlib import Path
from typing import Dict, List, Tuple, Any
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from collections import defaultdict

from db import get_all_desenhos


def find_header_row(sheet) -> int:
    """
    Find the row containing the table header (looks for "Nº." and "DESIGNAÇÃO").
    
    Returns:
        Row index (1-based) or None if not found
    """
    for row_idx in range(1, 20):  # Check first 20 rows
        row_values = [cell.value for cell in sheet[row_idx]]
        if "Nº." in row_values or "DESIGNAÇÃO" in row_values:
            return row_idx
    return None


def get_column_indices(sheet, header_row: int) -> Dict[str, int]:
    """
    Get column indices for key columns from header row.
    
    Returns:
        Dictionary mapping column name to column index (1-based)
    """
    columns = {}
    for col_idx, cell in enumerate(sheet[header_row], start=1):
        if cell.value:
            col_name = str(cell.value).strip()
            columns[col_name] = col_idx
    
    return columns


def find_elemento_anchors(sheet, header_row: int, col_indices: Dict[str, int]) -> List[Dict[str, Any]]:
    """
    Find all rows with ROW_KIND="ELEMENTO" which serve as anchors.
    
    Returns:
        List of anchor info: row_index, tipo_key, elemento_key
    """
    anchors = []
    
    row_kind_col = col_indices.get('ROW_KIND')
    tipo_key_col = col_indices.get('TIPO_KEY')
    elemento_key_col = col_indices.get('ELEMENTO_KEY')
    
    if not all([row_kind_col, tipo_key_col, elemento_key_col]):
        print("Warning: Missing required columns (ROW_KIND, TIPO_KEY, ELEMENTO_KEY)")
        return anchors
    
    # Start searching after header
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        row_kind = sheet.cell(row_idx, row_kind_col).value
        
        if row_kind == "ELEMENTO":
            tipo_key = sheet.cell(row_idx, tipo_key_col).value or ""
            elemento_key = sheet.cell(row_idx, elemento_key_col).value or ""
            
            anchors.append({
                'row_index': row_idx,
                'tipo_key': tipo_key,
                'elemento_key': elemento_key
            })
    
    return anchors


def delete_desenho_rows(sheet, anchor_row: int, tipo_key: str, elemento_key: str, col_indices: Dict[str, int]) -> int:
    """
    Delete all contiguous DESENHO rows below anchor with matching tipo_key/elemento_key.
    
    Returns:
        Number of rows deleted
    """
    row_kind_col = col_indices.get('ROW_KIND')
    tipo_key_col = col_indices.get('TIPO_KEY')
    elemento_key_col = col_indices.get('ELEMENTO_KEY')
    
    rows_to_delete = []
    current_row = anchor_row + 1
    
    # Find all contiguous DESENHO rows
    while current_row <= sheet.max_row:
        row_kind = sheet.cell(current_row, row_kind_col).value
        row_tipo = sheet.cell(current_row, tipo_key_col).value or ""
        row_elemento = sheet.cell(current_row, elemento_key_col).value or ""
        
        # Stop if we hit non-DESENHO or different tipo/elemento
        if row_kind != "DESENHO":
            break
        if row_tipo != tipo_key or row_elemento != elemento_key:
            break
        
        rows_to_delete.append(current_row)
        current_row += 1
    
    # Delete rows in reverse order to maintain indices
    for row_idx in reversed(rows_to_delete):
        sheet.delete_rows(row_idx, 1)
    
    return len(rows_to_delete)


def insert_desenho_rows(sheet, anchor_row: int, desenhos: List[Dict[str, Any]], col_indices: Dict[str, int]):
    """
    Insert desenho rows below anchor.
    """
    num_col = col_indices.get('Nº.')
    designacao_col = col_indices.get('DESIGNAÇÃO')
    ficheiro_col = col_indices.get('FICHEIRO')
    rev_col = col_indices.get('Rev')
    data_col = col_indices.get('DATA')
    row_kind_col = col_indices.get('ROW_KIND')
    tipo_key_col = col_indices.get('TIPO_KEY')
    elemento_key_col = col_indices.get('ELEMENTO_KEY')
    
    insert_row = anchor_row + 1
    
    for desenho in desenhos:
        # Insert new row
        sheet.insert_rows(insert_row, 1)
        
        # Populate cells
        if num_col:
            # Format: "ELEMENTO_KEY DES_NUM" (e.g., "FUN 01")
            num_value = f"{desenho['elemento_key']} {desenho['des_num']}" if desenho['elemento_key'] else desenho['des_num']
            sheet.cell(insert_row, num_col).value = num_value
        
        if designacao_col:
            # Use elemento_titulo or fall back to tipo_display
            designacao = desenho.get('elemento_titulo') or desenho.get('tipo_display', '')
            sheet.cell(insert_row, designacao_col).value = designacao
        
        if ficheiro_col:
            sheet.cell(insert_row, ficheiro_col).value = desenho['layout_name']
        
        if rev_col:
            sheet.cell(insert_row, rev_col).value = desenho.get('r', '')
        
        if data_col:
            sheet.cell(insert_row, data_col).value = desenho.get('data', '')
        
        if row_kind_col:
            sheet.cell(insert_row, row_kind_col).value = "DESENHO"
        
        if tipo_key_col:
            sheet.cell(insert_row, tipo_key_col).value = desenho['tipo_key']
        
        if elemento_key_col:
            sheet.cell(insert_row, elemento_key_col).value = desenho['elemento_key']
        
        insert_row += 1


def build_lpp_from_db(template_path: str, output_path: str, conn):
    """
    Generate LPP.xlsx from database using template.
    
    Args:
        template_path: Path to LPP_TEMPLATE.xlsx
        output_path: Path to output LPP.xlsx
        conn: Database connection
    """
    # Load template
    if not Path(template_path).exists():
        print(f"Error: Template not found at {template_path}")
        return
    
    wb = load_workbook(template_path)
    sheet = wb.active  # Assume first sheet
    
    # Find header row
    header_row = find_header_row(sheet)
    if not header_row:
        print("Error: Could not find header row with 'Nº.' and 'DESIGNAÇÃO'")
        return
    
    print(f"Header row found at: {header_row}")
    
    # Get column indices
    col_indices = get_column_indices(sheet, header_row)
    print(f"Columns found: {list(col_indices.keys())}")
    
    # Find ELEMENTO anchors
    anchors = find_elemento_anchors(sheet, header_row, col_indices)
    print(f"Found {len(anchors)} ELEMENTO anchors")
    
    # Get all desenhos from DB
    all_desenhos = get_all_desenhos(conn)
    print(f"Loaded {len(all_desenhos)} desenhos from database")
    
    # Group desenhos by (tipo_key, elemento_key)
    desenhos_by_key = defaultdict(list)
    for desenho in all_desenhos:
        key = (desenho['tipo_key'], desenho['elemento_key'])
        desenhos_by_key[key].append(desenho)
    
    # Process each anchor
    for anchor in anchors:
        tipo_key = anchor['tipo_key']
        elemento_key = anchor['elemento_key']
        row_index = anchor['row_index']
        
        print(f"\nProcessing anchor at row {row_index}: TIPO={tipo_key}, ELEMENTO={elemento_key}")
        
        # Delete existing DESENHO rows
        deleted = delete_desenho_rows(sheet, row_index, tipo_key, elemento_key, col_indices)
        print(f"  Deleted {deleted} existing rows")
        
        # Get desenhos for this anchor
        key = (tipo_key, elemento_key)
        desenhos = desenhos_by_key.get(key, [])
        
        if desenhos:
            # Insert new desenho rows
            insert_desenho_rows(sheet, row_index, desenhos, col_indices)
            print(f"  Inserted {len(desenhos)} new rows")
        else:
            print(f"  No desenhos found for this anchor")
    
    # Save output
    output_path_obj = Path(output_path)
    output_path_obj.parent.mkdir(parents=True, exist_ok=True)
    
    wb.save(output_path)
    print(f"\nLPP saved to: {output_path}")
